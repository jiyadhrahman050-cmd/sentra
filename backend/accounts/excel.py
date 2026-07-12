from openpyxl import Workbook
from django.http import HttpResponse


def export_to_excel(filename, sheet_name, headers, queryset, row_builder):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    # Header
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column).value = header

    # Data
    row_number = 2

    for obj in queryset:
        values = row_builder(obj)

        for column, value in enumerate(values, start=1):
            worksheet.cell(
                row=row_number,
                column=column
            ).value = value

        row_number += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}.xlsx"'
    )

    workbook.save(response)

    return response